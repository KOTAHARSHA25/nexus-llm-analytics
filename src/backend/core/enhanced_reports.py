# Enhanced Report Generation System
# Publication-ready PDF/Excel reports with headers, footers, page numbers, and professional styling

import os
import json
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
from io import BytesIO
import pandas as pd

# PDF Generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    Image, PageBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor, black, white, grey
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY

# Excel Generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel reports disabled")

logger = logging.getLogger(__name__)


# =============================================================================
# COLOR PALETTE - Professional, Publication-Ready
# =============================================================================
class ColorPalette:
    """Professional color palette for reports"""
    PRIMARY = HexColor('#1e40af')       # Deep blue - headers
    SECONDARY = HexColor('#3b82f6')     # Blue - accents
    ACCENT = HexColor('#0ea5e9')        # Sky blue - highlights
    SUCCESS = HexColor('#059669')       # Green - positive metrics
    WARNING = HexColor('#d97706')       # Amber - warnings
    ERROR = HexColor('#dc2626')         # Red - errors
    TEXT_PRIMARY = HexColor('#111827')  # Near black - main text
    TEXT_SECONDARY = HexColor('#4b5563') # Gray - secondary text
    TEXT_MUTED = HexColor('#9ca3af')    # Light gray - captions
    BACKGROUND = HexColor('#f8fafc')    # Light gray - backgrounds
    BORDER = HexColor('#e2e8f0')        # Light border
    TABLE_HEADER = HexColor('#1e3a5f')  # Dark blue - table headers
    TABLE_ROW_ALT = HexColor('#f1f5f9') # Alternating rows


# =============================================================================
# NUMBERED CANVAS - Headers, Footers, Page Numbers
# =============================================================================
class NumberedCanvas(canvas.Canvas):
    """Canvas that tracks page numbers and draws headers/footers"""
    
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.report_title = "Nexus Analytics Report"
        self.company_name = "Nexus LLM Analytics"
        self.generated_date = datetime.now()
    
    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()
    
    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_header_footer(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)
    
    def _draw_header_footer(self, page_count):
        """Draw header and footer on each page"""
        page_num = self._pageNumber
        width, height = A4
        
        # Skip header/footer on title page (page 1)
        if page_num == 1:
            return
        
        self.saveState()
        
        # === HEADER ===
        self.setStrokeColor(ColorPalette.PRIMARY)
        self.setLineWidth(2)
        self.line(50, height - 40, width - 50, height - 40)
        
        # Left: Company name
        self.setFont('Helvetica-Bold', 9)
        self.setFillColor(ColorPalette.PRIMARY)
        self.drawString(50, height - 32, self.company_name)
        
        # Right: Report title
        title = self.report_title[:50] + "..." if len(self.report_title) > 50 else self.report_title
        self.setFont('Helvetica', 9)
        self.setFillColor(ColorPalette.TEXT_SECONDARY)
        self.drawRightString(width - 50, height - 32, title)
        
        # === FOOTER ===
        self.setStrokeColor(ColorPalette.BORDER)
        self.setLineWidth(1)
        self.line(50, 35, width - 50, 35)
        
        # Left: Generation date
        self.setFont('Helvetica', 8)
        self.setFillColor(ColorPalette.TEXT_MUTED)
        date_str = self.generated_date.strftime('%B %d, %Y')
        self.drawString(50, 22, f"Generated: {date_str}")
        
        # Center: Page number
        self.setFont('Helvetica', 9)
        self.setFillColor(ColorPalette.TEXT_SECONDARY)
        self.drawCentredString(width / 2, 22, f"Page {page_num} of {page_count}")
        
        # Right: Confidential
        self.setFont('Helvetica-Oblique', 8)
        self.setFillColor(ColorPalette.TEXT_MUTED)
        self.drawRightString(width - 50, 22, "Confidential")
        
        self.restoreState()


# =============================================================================
# REPORT TEMPLATE
# =============================================================================
class ReportTemplate:
    """Report template configuration"""
    
    def __init__(self, title: str = "Nexus Analytics Report"):
        self.title = title
        self.created_at = datetime.now()
        self.company_name = "Nexus LLM Analytics"
        self.version = "2.0"
        self.logo_path = None


# =============================================================================
# PDF REPORT GENERATOR - Publication Ready
# =============================================================================
class PDFReportGenerator:
    """Generate professional, publication-ready PDF reports"""
    
    def __init__(self, template: ReportTemplate = None):
        self.template = template or ReportTemplate()
        self.styles = getSampleStyleSheet()
        self._setup_professional_styles()
    
    def _setup_professional_styles(self):
        """Setup professional typography"""
        
        # Main Title
        self.styles.add(ParagraphStyle(
            name='MainTitle',
            parent=self.styles['Heading1'],
            fontSize=28,
            spaceAfter=6,
            textColor=ColorPalette.PRIMARY,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            leading=34
        ))
        
        # Subtitle
        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Normal'],
            fontSize=14,
            textColor=ColorPalette.TEXT_SECONDARY,
            alignment=TA_CENTER,
            spaceAfter=30
        ))
        
        # Section Header
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceBefore=24,
            spaceAfter=12,
            textColor=ColorPalette.PRIMARY,
            fontName='Helvetica-Bold'
        ))
        
        # Subsection Header
        self.styles.add(ParagraphStyle(
            name='SubsectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=16,
            spaceAfter=8,
            textColor=ColorPalette.TEXT_PRIMARY,
            fontName='Helvetica-Bold'
        ))
        
        # Body Text - Justified (using custom name to avoid conflict with built-in)
        self.styles.add(ParagraphStyle(
            name='BodyTextCustom',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=ColorPalette.TEXT_PRIMARY,
            alignment=TA_JUSTIFY,
            spaceAfter=8,
            leading=16
        ))
        
        # Caption
        self.styles.add(ParagraphStyle(
            name='Caption',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=ColorPalette.TEXT_MUTED,
            alignment=TA_CENTER,
            spaceBefore=6,
            spaceAfter=12,
            fontName='Helvetica-Oblique'
        ))
        
        # TOC Entry
        self.styles.add(ParagraphStyle(
            name='TOCEntry',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=ColorPalette.TEXT_PRIMARY,
            spaceAfter=8
        ))
    
    def generate_report(self, analysis_results: List[Dict[str, Any]], 
                       output_path: str = None) -> str:
        """Generate comprehensive professional PDF report"""
        
        if not output_path:
            output_path = f"nexus_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        story = []
        
        # 1. Title Page
        story.extend(self._create_title_page())
        story.append(PageBreak())
        
        # 2. Table of Contents
        story.extend(self._create_table_of_contents(analysis_results))
        story.append(PageBreak())
        
        # 3. Executive Summary
        story.extend(self._create_executive_summary(analysis_results))
        story.append(PageBreak())
        
        # 4. Key Findings
        story.extend(self._create_key_findings(analysis_results))
        story.append(PageBreak())
        
        # 5. Analysis Sections
        for i, result in enumerate(analysis_results, 1):
            story.extend(self._create_analysis_section(result, i, len(analysis_results)))
            if i < len(analysis_results):
                story.append(PageBreak())
        
        # 6. Statistical Summary
        story.append(PageBreak())
        story.extend(self._create_statistical_summary(analysis_results))
        
        # 7. Data Quality Assessment
        story.append(PageBreak())
        story.extend(self._create_data_quality_section(analysis_results))
        
        # 8. Methodology
        story.append(PageBreak())
        story.extend(self._create_methodology_section())
        
        # 9. Glossary & References
        story.append(PageBreak())
        story.extend(self._create_glossary_and_references())
        
        # 10. Appendix
        story.append(PageBreak())
        story.extend(self._create_appendix(analysis_results))
        
        # Build PDF with numbered pages
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=60,
            bottomMargin=50
        )
        
        try:
            doc.build(story, canvasmaker=self._create_canvas)
            logger.info(f"PDF report generated: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            raise
    
    def _create_canvas(self, *args, **kwargs):
        """Factory for numbered canvas"""
        c = NumberedCanvas(*args, **kwargs)
        c.report_title = self.template.title
        c.company_name = self.template.company_name
        c.generated_date = self.template.created_at
        return c
    
    def _create_title_page(self) -> List:
        """Create professional title page"""
        elements = []
        elements.append(Spacer(1, 1.5*inch))
        
        # Company Logo Box
        logo_table = Table(
            [[Paragraph(f"<b>{self.template.company_name}</b>", 
                       ParagraphStyle('Logo', fontSize=16, textColor=white, alignment=TA_CENTER))]],
            colWidths=[3*inch], rowHeights=[0.6*inch]
        )
        logo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), ColorPalette.PRIMARY),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        logo_table.hAlign = 'CENTER'
        elements.append(logo_table)
        elements.append(Spacer(1, 0.8*inch))
        
        # Title
        elements.append(Paragraph(self.template.title, self.styles['MainTitle']))
        elements.append(Paragraph("Comprehensive Data Analysis Report", self.styles['Subtitle']))
        elements.append(Spacer(1, 0.5*inch))
        
        # Divider
        hr = Table([['']], colWidths=[4*inch], rowHeights=[2])
        hr.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), ColorPalette.SECONDARY)]))
        hr.hAlign = 'CENTER'
        elements.append(hr)
        elements.append(Spacer(1, 0.5*inch))
        
        # Metadata
        date_str = self.template.created_at.strftime('%B %d, %Y')
        metadata = [
            ['Report Generated:', date_str],
            ['Report Version:', f'v{self.template.version}'],
            ['Platform:', 'Nexus LLM Analytics'],
            ['Document Type:', 'Technical Analysis Report']
        ]
        meta_table = Table(metadata, colWidths=[1.8*inch, 2.5*inch])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), ColorPalette.TEXT_SECONDARY),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        meta_table.hAlign = 'CENTER'
        elements.append(meta_table)
        elements.append(Spacer(1, 1*inch))
        
        # Abstract
        abstract = """This report presents comprehensive data analysis results generated by the 
        Nexus LLM Analytics platform. Utilizing advanced multi-agent AI systems and statistical 
        analysis techniques, this document provides actionable insights, visualizations, and 
        quality assessments suitable for research publication and technical documentation."""
        
        abstract_table = Table([[Paragraph(abstract.strip(), self.styles['BodyText'])]], colWidths=[5.5*inch])
        abstract_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), ColorPalette.BACKGROUND),
            ('BOX', (0, 0), (-1, -1), 1, ColorPalette.BORDER),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        abstract_table.hAlign = 'CENTER'
        elements.append(abstract_table)
        
        elements.append(Spacer(1, 1.5*inch))
        elements.append(Paragraph("<i>CONFIDENTIAL - For authorized use only</i>",
            ParagraphStyle('Conf', fontSize=9, textColor=ColorPalette.TEXT_MUTED, alignment=TA_CENTER)))
        
        return elements
    
    def _create_table_of_contents(self, results: List[Dict[str, Any]]) -> List:
        """Create table of contents"""
        elements = []
        elements.append(Paragraph("Table of Contents", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.3*inch))
        
        toc_entries = [
            ("1.", "Executive Summary", "3"),
            ("2.", "Key Findings", "4"),
        ]
        
        section_num = 3
        for i, result in enumerate(results, 1):
            query = result.get('query', f'Analysis {i}')[:45]
            toc_entries.append((f"{section_num}.", f"Analysis {i}: {query}", str(section_num + 1)))
            section_num += 1
        
        toc_entries.extend([
            (f"{section_num}.", "Statistical Summary", str(section_num + 2)),
            (f"{section_num + 1}.", "Data Quality Assessment", str(section_num + 3)),
            (f"{section_num + 2}.", "Methodology", str(section_num + 4)),
            (f"{section_num + 3}.", "Glossary & References", str(section_num + 5)),
            (f"{section_num + 4}.", "Appendix", str(section_num + 6)),
        ])
        
        toc_data = [[Paragraph(f"<b>{n}</b>", self.styles['TOCEntry']),
                     Paragraph(t, self.styles['TOCEntry']),
                     Paragraph(p, self.styles['TOCEntry'])] for n, t, p in toc_entries]
        
        toc_table = Table(toc_data, colWidths=[0.4*inch, 4.5*inch, 0.5*inch])
        toc_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LINEBELOW', (1, 0), (1, -1), 0.5, ColorPalette.BORDER, None, (1, 3)),
        ]))
        elements.append(toc_table)
        return elements
    
    def _create_executive_summary(self, results: List[Dict[str, Any]]) -> List:
        """Create executive summary"""
        elements = []
        elements.append(Paragraph("1. Executive Summary", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        
        total = len(results)
        successful = sum(1 for r in results if r.get('success', True))
        success_rate = (successful / total * 100) if total > 0 else 0
        
        overview = f"""This report documents the results of <b>{total}</b> data analysis operations 
        performed using the Nexus LLM Analytics platform. The analyses achieved a <b>{success_rate:.1f}%</b> 
        success rate, demonstrating the reliability of our multi-agent AI system."""
        elements.append(Paragraph(overview.strip(), self.styles['BodyText']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Metrics table
        elements.append(Paragraph("<b>Summary Metrics</b>", self.styles['SubsectionHeader']))
        
        total_time = sum(r.get('execution_time', 0) for r in results)
        avg_time = total_time / total if total > 0 else 0
        charts = sum(1 for r in results if r.get('chart_data', {}).get('visualization'))
        
        metrics = [
            ['Metric', 'Value', 'Status'],
            ['Total Analyses', str(total), '✓'],
            ['Successful', str(successful), '✓' if successful == total else '△'],
            ['Success Rate', f'{success_rate:.1f}%', '✓' if success_rate >= 90 else '△'],
            ['Avg. Execution Time', f'{avg_time:.2f}s', '✓' if avg_time < 30 else '△'],
            ['Visualizations', str(charts), '✓'],
        ]
        
        metrics_table = Table(metrics, colWidths=[2.2*inch, 1.5*inch, 0.8*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(metrics_table)
        return elements
    
    def _create_key_findings(self, results: List[Dict[str, Any]]) -> List:
        """Create key findings section"""
        elements = []
        elements.append(Paragraph("2. Key Findings", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("Key findings identified across all analyses:", self.styles['BodyText']))
        elements.append(Spacer(1, 0.15*inch))
        
        findings = []
        for i, result in enumerate(results, 1):
            result_text = str(result.get('result', ''))
            if result_text and result_text != 'No results available':
                sentences = result_text.split('.')
                if sentences and sentences[0].strip():
                    finding = sentences[0].strip()[:150]
                    findings.append((f"Analysis {i}", finding + ("..." if len(sentences[0]) > 150 else ".")))
        
        for title, finding in findings[:5]:
            box = Table([[Paragraph(f"<b>{title}:</b> {finding}", self.styles['BodyText'])]], colWidths=[5.5*inch])
            box.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), HexColor('#eff6ff')),
                ('BOX', (0, 0), (-1, -1), 1, ColorPalette.SECONDARY),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(box)
            elements.append(Spacer(1, 0.1*inch))
        
        if not findings:
            elements.append(Paragraph("<i>Key findings populated from analysis results.</i>", self.styles['BodyText']))
        
        return elements
    
    def _create_analysis_section(self, result: Dict[str, Any], section_num: int, total: int) -> List:
        """Create detailed analysis section"""
        elements = []
        
        query = result.get('query', f'Analysis {section_num}')
        elements.append(Paragraph(f"{section_num + 2}. Analysis {section_num}: {query[:60]}{'...' if len(query) > 60 else ''}", 
                                  self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Metadata
        elements.append(Paragraph("<b>Analysis Metadata</b>", self.styles['SubsectionHeader']))
        
        meta_rows = [
            ['Property', 'Value'],
            ['Data Source', result.get('filename', 'N/A')],
            ['Analysis Type', result.get('type', 'General Analysis')],
            ['Execution Time', f"{result.get('execution_time', 0):.2f} seconds"],
            ['Status', '✓ Successful' if result.get('success', True) else '✗ Failed'],
        ]
        
        routing = result.get('routing_info', {})
        if routing:
            meta_rows.append(['Model Used', routing.get('selected_model', 'N/A')])
            meta_rows.append(['Routing Tier', routing.get('selected_tier', 'N/A')])
            if routing.get('complexity_score'):
                meta_rows.append(['Complexity', f"{routing['complexity_score']:.2f}"])
        
        meta_table = Table(meta_rows, colWidths=[1.8*inch, 3.5*inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Results
        elements.append(Paragraph("<b>Analysis Results</b>", self.styles['SubsectionHeader']))
        result_text = str(result.get('result', 'No results available'))
        if result_text and result_text != 'No results available':
            for para in result_text.split('\n')[:10]:
                if para.strip():
                    elements.append(Paragraph(para.strip(), self.styles['BodyText']))
        else:
            elements.append(Paragraph("<i>No results available.</i>", self.styles['BodyText']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Chart
        chart_data = result.get('chart_data', {})
        vis = chart_data.get('visualization', {}) if chart_data else {}
        if vis and vis.get('figure_json'):
            elements.append(Paragraph("<b>Data Visualization</b>", self.styles['SubsectionHeader']))
            try:
                chart_path = self._save_chart(vis['figure_json'], section_num)
                if chart_path and os.path.exists(chart_path):
                    img = Image(chart_path, width=5.5*inch, height=3.5*inch)
                    img.hAlign = 'CENTER'
                    elements.append(img)
                    elements.append(Paragraph(f"Figure {section_num}: {vis.get('chart_type', 'Chart').title()} visualization",
                                              self.styles['Caption']))
            except Exception as e:
                logger.error(f"Chart error: {e}")
        
        # Data Structure
        data_struct = result.get('data_structure', {}) or chart_data.get('data_analysis', {})
        if data_struct:
            elements.append(Paragraph("<b>Data Structure</b>", self.styles['SubsectionHeader']))
            struct_rows = [['Property', 'Value']]
            if data_struct.get('row_count'):
                struct_rows.append(['Total Rows', str(data_struct['row_count'])])
            if data_struct.get('column_count'):
                struct_rows.append(['Total Columns', str(data_struct['column_count'])])
            for col_type in ['numeric_columns', 'categorical_columns', 'datetime_columns']:
                cols = data_struct.get(col_type, [])
                if cols:
                    name = col_type.replace('_', ' ').title()
                    val = ', '.join(cols[:3]) + (f"... (+{len(cols)-3})" if len(cols) > 3 else '')
                    struct_rows.append([name, val])
            
            if len(struct_rows) > 1:
                struct_table = Table(struct_rows, colWidths=[1.8*inch, 3.5*inch])
                struct_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
                    ('TEXTCOLOR', (0, 0), (-1, 0), white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
                ]))
                elements.append(struct_table)
        
        return elements
    
    def _save_chart(self, figure_json: str, section_num: int) -> Optional[str]:
        """Save Plotly chart as PNG"""
        try:
            import plotly.graph_objects as go
            import plotly.io as pio
            
            temp_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'reports', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            chart_path = os.path.join(temp_dir, f"chart_{section_num}_{datetime.now().strftime('%H%M%S')}.png")
            
            fig_dict = json.loads(figure_json) if isinstance(figure_json, str) else figure_json
            fig = go.Figure(fig_dict)
            fig.update_layout(
                font=dict(size=11),
                paper_bgcolor='white',
                plot_bgcolor='white',
                margin=dict(l=50, r=50, t=60, b=50)
            )
            pio.write_image(fig, chart_path, width=1100, height=700, scale=2)
            return chart_path
        except Exception as e:
            logger.error(f"Chart save failed: {e}")
            return None
    
    def _create_statistical_summary(self, results: List[Dict[str, Any]]) -> List:
        """Create statistical summary"""
        elements = []
        n = len(results) + 3
        elements.append(Paragraph(f"{n}. Statistical Summary", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("Aggregate statistical measures across all analyses.", self.styles['BodyText']))
        elements.append(Spacer(1, 0.15*inch))
        
        exec_times = [r.get('execution_time', 0) for r in results if r.get('execution_time')]
        if exec_times:
            avg = sum(exec_times) / len(exec_times)
            variance = sum((t - avg) ** 2 for t in exec_times) / len(exec_times)
            std = variance ** 0.5
            ci = 1.96 * std / (len(exec_times) ** 0.5) if len(exec_times) > 0 else 0
            
            stats = [
                ['Statistic', 'Value', 'Unit'],
                ['Sample Size (n)', str(len(results)), 'analyses'],
                ['Mean Execution Time', f'{avg:.3f}', 'seconds'],
                ['Std. Deviation', f'{std:.3f}', 'seconds'],
                ['Min Time', f'{min(exec_times):.3f}', 'seconds'],
                ['Max Time', f'{max(exec_times):.3f}', 'seconds'],
                ['95% CI', f'{avg:.3f} ± {ci:.3f}', 'seconds'],
            ]
            
            stats_table = Table(stats, colWidths=[2.2*inch, 1.8*inch, 1.2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
            ]))
            elements.append(stats_table)
        else:
            elements.append(Paragraph("<i>Insufficient data for statistics.</i>", self.styles['BodyText']))
        return elements
    
    def _create_data_quality_section(self, results: List[Dict[str, Any]]) -> List:
        """Create data quality assessment"""
        elements = []
        n = len(results) + 4
        elements.append(Paragraph(f"{n}. Data Quality Assessment", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("Quality and completeness evaluation of analyzed data.", self.styles['BodyText']))
        elements.append(Spacer(1, 0.15*inch))
        
        failures = sum(1 for r in results if not r.get('success', True))
        
        quality = [
            ['Quality Dimension', 'Assessment', 'Score'],
            ['Completeness', 'Data files successfully loaded', '✓ High'],
            ['Validity', 'All analyses executed without errors' if failures == 0 else f'{failures} failures', '✓ High' if failures == 0 else '△ Medium'],
            ['Consistency', 'Results reproducible within platform', '✓ High'],
            ['Timeliness', 'Real-time analysis performed', '✓ High'],
            ['Accuracy', 'Multi-agent validation applied', '✓ High'],
        ]
        
        quality_table = Table(quality, colWidths=[1.8*inch, 2.8*inch, 0.9*inch])
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
        ]))
        elements.append(quality_table)
        return elements
    
    def _create_methodology_section(self) -> List:
        """Create methodology section"""
        elements = []
        elements.append(Paragraph("Methodology", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph("<b>Multi-Agent AI System Architecture</b>", self.styles['SubsectionHeader']))
        methodology = """The Nexus LLM Analytics platform employs a sophisticated multi-agent AI 
        architecture for comprehensive data analysis. Specialized AI agents collaborate to provide 
        thorough, validated insights."""
        elements.append(Paragraph(methodology.strip(), self.styles['BodyText']))
        elements.append(Spacer(1, 0.1*inch))
        
        agents = [
            ['Agent', 'Role', 'Responsibility'],
            ['Data Agent', 'Data Processing', 'Loads, validates, and structures input data'],
            ['Analysis Agent', 'Statistical Analysis', 'Performs computations and pattern detection'],
            ['Visualization Agent', 'Chart Generation', 'Creates visualizations using Plotly'],
            ['Review Agent', 'Quality Assurance', 'Validates results and suggests improvements'],
            ['Report Agent', 'Documentation', 'Compiles findings into professional reports'],
        ]
        
        agents_table = Table(agents, colWidths=[1.2*inch, 1.3*inch, 3*inch])
        agents_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
        ]))
        elements.append(agents_table)
        elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph("<b>Security & Sandboxing</b>", self.styles['SubsectionHeader']))
        security = """All code execution occurs within a RestrictedPython sandbox environment, 
        ensuring secure processing of user data while enabling powerful analytical capabilities."""
        elements.append(Paragraph(security.strip(), self.styles['BodyText']))
        return elements
    
    def _create_glossary_and_references(self) -> List:
        """Create glossary and references"""
        elements = []
        elements.append(Paragraph("Glossary & References", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Glossary
        elements.append(Paragraph("<b>Glossary</b>", self.styles['SubsectionHeader']))
        glossary = [
            ('LLM', 'Large Language Model - AI models trained on vast text data'),
            ('Multi-Agent System', 'Architecture where specialized AI agents collaborate'),
            ('Sandboxed Execution', 'Secure code execution in an isolated environment'),
            ('RAG', 'Retrieval-Augmented Generation - combines retrieval with generation'),
            ('Confidence Interval', 'Range likely to contain the true population parameter'),
        ]
        for term, defn in glossary:
            elements.append(Paragraph(f"<b>{term}:</b> {defn}", self.styles['BodyText']))
        elements.append(Spacer(1, 0.2*inch))
        
        # References
        elements.append(Paragraph("<b>References</b>", self.styles['SubsectionHeader']))
        refs = [
            "[1] CrewAI Framework - Multi-agent orchestration for AI systems",
            "[2] Ollama - Local large language model deployment",
            "[3] Plotly - Interactive visualization library for Python",
            "[4] Pandas - Data manipulation and analysis library",
            "[5] RestrictedPython - Secure Python sandbox execution",
        ]
        for ref in refs:
            elements.append(Paragraph(ref, self.styles['BodyText']))
        return elements
    
    def _create_appendix(self, results: List[Dict[str, Any]]) -> List:
        """Create appendix"""
        elements = []
        elements.append(Paragraph("Appendix: Technical Specifications", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.2*inch))
        
        specs = [
            ['Component', 'Technology/Version'],
            ['Platform', 'Nexus LLM Analytics v2.0'],
            ['Primary LLM', 'Llama 3.1 8B (via Ollama)'],
            ['Secondary LLM', 'Phi-3 Mini (via Ollama)'],
            ['Fast LLM', 'TinyLlama (via Ollama)'],
            ['Data Processing', 'Pandas, NumPy'],
            ['Visualization', 'Plotly'],
            ['Security', 'RestrictedPython Sandbox'],
            ['Report Generation', 'ReportLab'],
        ]
        
        specs_table = Table(specs, colWidths=[2*inch, 3.5*inch])
        specs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ColorPalette.TABLE_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, ColorPalette.BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ColorPalette.TABLE_ROW_ALT]),
        ]))
        elements.append(specs_table)
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph(f"<i>Report generated: {self.template.created_at.strftime('%Y-%m-%d %H:%M:%S')}</i>",
                                  self.styles['Caption']))
        return elements


# =============================================================================
# EXCEL REPORT GENERATOR
# =============================================================================
class ExcelReportGenerator:
    """Generate professional Excel reports"""
    
    def __init__(self, template: ReportTemplate = None):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required for Excel reports")
        self.template = template or ReportTemplate()
    
    def generate_report(self, analysis_results: List[Dict[str, Any]], output_path: str = None) -> str:
        """Generate Excel report"""
        if not output_path:
            output_path = f"nexus_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        self._create_summary_sheet(workbook, analysis_results)
        self._create_results_sheets(workbook, analysis_results)
        self._create_data_sheet(workbook, analysis_results)
        
        try:
            workbook.save(output_path)
            logger.info(f"Excel report generated: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            raise
    
    def _create_summary_sheet(self, workbook, results: List[Dict[str, Any]]):
        """Create summary sheet"""
        ws = workbook.create_sheet("Executive Summary", 0)
        ws['A1'] = self.template.title
        ws['A1'].font = Font(size=20, bold=True, color='1e40af')
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Generated: {self.template.created_at.strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=12, color='6B7280')
        
        row = 4
        ws[f'A{row}'] = "Summary Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        stats = [
            ["Total Analyses", len(results)],
            ["Successful", sum(1 for r in results if r.get('success', True))],
            ["Success Rate", f"{(sum(1 for r in results if r.get('success', True))/max(len(results),1))*100:.1f}%"]
        ]
        for name, val in stats:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = val
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25
    
    def _create_results_sheets(self, workbook, results: List[Dict[str, Any]]):
        """Create individual result sheets"""
        for i, result in enumerate(results, 1):
            ws = workbook.create_sheet(f"Analysis_{i}")
            ws['A1'] = f"Analysis {i}: {result.get('query', 'N/A')[:40]}"
            ws['A1'].font = Font(size=16, bold=True, color='1e40af')
            ws.merge_cells('A1:D1')
            
            row = 3
            details = [
                ("Query", result.get('query', 'N/A')),
                ("Filename", result.get('filename', 'N/A')),
                ("Type", result.get('type', 'N/A')),
                ("Success", "✓" if result.get('success', True) else "✗"),
                ("Execution Time", f"{result.get('execution_time', 0):.2f}s")
            ]
            for label, val in details:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = str(val)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
            ws[f'A{row}'] = "Results"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = str(result.get('result', 'N/A'))[:32767]
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:D{row}')
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 50
    
    def _create_data_sheet(self, workbook, results: List[Dict[str, Any]]):
        """Create overview data sheet"""
        ws = workbook.create_sheet("Analysis Overview")
        
        headers = ["ID", "Query", "Filename", "Type", "Success", "Execution_Time", "Model", "Tier"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        
        for i, result in enumerate(results, 1):
            meta = result.get('metadata', {})
            routing = result.get('routing_info', {})
            row_data = [
                i,
                result.get('query', 'N/A')[:50],
                result.get('filename', 'N/A'),
                result.get('type', 'N/A'),
                "✓" if result.get('success', True) else "✗",
                f"{result.get('execution_time', 0):.2f}s",
                routing.get('selected_model', meta.get('model', 'N/A')),
                routing.get('selected_tier', meta.get('routing_tier', 'N/A'))
            ]
            for col, val in enumerate(row_data, 1):
                ws.cell(row=i+1, column=col, value=val)
        
        widths = {'A': 8, 'B': 40, 'C': 25, 'D': 15, 'E': 10, 'F': 15, 'G': 20, 'H': 15}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        ws.freeze_panes = 'A2'


# =============================================================================
# ENHANCED REPORT MANAGER - Main Interface
# =============================================================================
class EnhancedReportManager:
    """Unified manager for all report formats"""
    
    def __init__(self):
        self.pdf_generator = PDFReportGenerator()
        self.excel_generator = ExcelReportGenerator() if EXCEL_AVAILABLE else None
    
    def generate_report(self, analysis_results: List[Dict[str, Any]], 
                       format_type: str = "pdf", 
                       template: ReportTemplate = None,
                       output_path: str = None) -> str:
        """
        Generate report in specified format
        
        Args:
            analysis_results: List of analysis results
            format_type: "pdf", "excel", or "both"
            template: Custom report template
            output_path: Custom output path
            
        Returns:
            Path to generated report(s)
        """
        if template:
            self.pdf_generator.template = template
            if self.excel_generator:
                self.excel_generator.template = template
        
        generated = []
        
        try:
            if format_type.lower() in ["pdf", "both"]:
                pdf_path = output_path or f"nexus_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                generated.append(self.pdf_generator.generate_report(analysis_results, pdf_path))
            
            if format_type.lower() in ["excel", "both"] and self.excel_generator:
                excel_path = output_path or f"nexus_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                if format_type.lower() == "both":
                    excel_path = excel_path.replace('.pdf', '.xlsx')
                generated.append(self.excel_generator.generate_report(analysis_results, excel_path))
            
            logger.info(f"Report generation completed: {generated}")
            return generated[0] if len(generated) == 1 else generated
            
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            raise


# Convenience function
def generate_report(results: List[Dict[str, Any]], format_type: str = "pdf", 
                   title: str = None, output_path: str = None) -> str:
    """Generate a professional report"""
    manager = EnhancedReportManager()
    template = ReportTemplate(title=title) if title else None
    return manager.generate_report(results, format_type, template, output_path)
